from flask import Flask, request, jsonify
from flask_cors import CORS
import cv2
import numpy as np
import face_recognition
import os
from datetime import datetime
import base64
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app)

# Create directory for known faces
if not os.path.exists('known_faces'):
    os.makedirs('known_faces')

# Excel file setup
EXCEL_FILE = 'attendance_records.xlsx'

def initialize_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance"
        headers = ["Name", "Date", "Time", "Status"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header
        wb.save(EXCEL_FILE)

initialize_excel()

def save_to_excel(name, timestamp, status):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    # Split timestamp into date and time
    dt = datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
    date_str = dt.strftime("%Y-%m-%d")
    time_str = dt.strftime("%H:%M:%S")
    
    # Append new row
    ws.append([name, date_str, time_str, status])
    wb.save(EXCEL_FILE)

# Load known faces
known_face_encodings = []
known_face_names = []

def load_known_faces():
    global known_face_encodings, known_face_names
    known_face_encodings = []
    known_face_names = []
    
    for filename in os.listdir('known_faces'):
        if filename.endswith('.jpg') or filename.endswith('.png'):
            image = face_recognition.load_image_file(f'known_faces/{filename}')
            encoding = face_recognition.face_encodings(image)
            if encoding:
                known_face_encodings.append(encoding[0])
                known_face_names.append(os.path.splitext(filename)[0])

load_known_faces()

# Attendance log
attendance_log = []

@app.route('/mark-attendance', methods=['POST'])
def mark_attendance():
    data = request.json
    image_data = data['image'].split(',')[1]
    image_bytes = base64.b64decode(image_data)
    image_array = np.frombuffer(image_bytes, dtype=np.uint8)
    frame = cv2.imdecode(image_array, cv2.IMREAD_COLOR)
    rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    face_locations = face_recognition.face_locations(rgb_frame)
    face_encodings = face_recognition.face_encodings(rgb_frame, face_locations)
    
    if not face_encodings:
        return jsonify({"error": "No face detected"})

    for face_encoding in face_encodings:
        matches = face_recognition.compare_faces(known_face_encodings, face_encoding, tolerance=0.6)
        name = "Unknown"
        
        if True in matches:
            first_match_index = matches.index(True)
            name = known_face_names[first_match_index]
            
            today = datetime.now().strftime("%Y-%m-%d")
            already_marked = any(
                entry["name"] == name and entry["timestamp"].startswith(today)
                for entry in attendance_log
            )
            
            if already_marked:
                return jsonify({"error": f"{name} has already marked attendance today"})
            
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            attendance_log.append({"name": name, "timestamp": timestamp, "status": "Present"})
            
            # Save to Excel file
            save_to_excel(name, timestamp, "Present")
            
            return jsonify({"name": name, "status": "Present"})
    
    return jsonify({"error": "No face matched"})

@app.route('/add-face', methods=['POST'])
def add_face():
    data = request.json
    name = data['name']
    image_data = data['image'].split(',')[1]
    image_bytes = base64.b64decode(image_data)
    
    with open(f'known_faces/{name}.jpg', 'wb') as f:
        f.write(image_bytes)
    
    load_known_faces()
    return jsonify({"success": True})

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)